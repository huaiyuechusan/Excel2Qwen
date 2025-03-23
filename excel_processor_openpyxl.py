import os
from openpyxl import load_workbook

def process_excel_files(directory_path):
    """
    使用openpyxl处理指定目录下的所有Excel文件，提取每个文件所有sheet的B列数据
    
    参数:
        directory_path (str): Excel文件所在的目录路径
        
    返回:
        dict: 包含所有sheet的B列数据，键为sheet名称
    """
    # 初始化一个字典，用于存储所有sheet的B列数据
    all_sheets_data = {}
    
    # 获取目录中所有Excel文件
    excel_files = [f for f in os.listdir(directory_path) 
                  if f.endswith('.xlsx') or f.endswith('.xlsm')]
    
    # 遍历所有Excel文件
    for file_name in excel_files:
        file_path = os.path.join(directory_path, file_name)
        
        try:
            # 加载工作簿
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # 遍历所有sheet
            for sheet_name in wb.sheetnames:
                # 如果字典中还没有这个sheet的键，则创建一个空列表
                if sheet_name not in all_sheets_data:
                    all_sheets_data[sheet_name] = []
                
                # 处理当前sheet
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第2行开始，跳过标题行
                    if len(row) > 1 and row[1] is not None:  # 确保B列有值
                        all_sheets_data[sheet_name].append(row[1])
            
            # 关闭工作簿
            wb.close()
                
        except Exception as e:
            print(f"处理文件 {file_name} 时出错: {str(e)}")
    
    return all_sheets_data

# 使用示例
if __name__ == "__main__":
    directory = "your_excel_files_directory"  # 替换为实际的目录路径
    all_sheets_data = process_excel_files(directory)
    
    # 打印每个sheet的B列数据
    for sheet_name, data_list in all_sheets_data.items():
        print(f"\n{sheet_name} 工作表的B列数据:")
        print(data_list) 